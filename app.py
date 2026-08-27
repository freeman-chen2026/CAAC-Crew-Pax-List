import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import re
from datetime import datetime, timedelta
import traceback
import json
import os

# ---------- 页面设置 ----------
st.set_page_config(page_title="备案表&世界时行程&航路处理", layout="wide")
st.title("🛫 备案表 / 世界时行程 / 航路处理")

# ---------- 创建选项卡 ----------
tab1, tab2, tab3 = st.tabs(["📋 功能1：备案表生成", "🌐 功能2：世界时行程", "✈️ 功能3：航路处理工具"])

# ================================================================
# 功能1：备案表生成
# ================================================================
with tab1:
    st.markdown("上传 GD单 和模板，自动生成备案表（联系方式、执照号码及证件号码已内置）。")

    # ---------- 内置联系方式映射 ----------
    BUILTIN_CONTACT_MAP = {
        "庚凡": "139 2463 9747",
        "张永一": "139 0125 9544",
        "梅峰": "135 0967 8127",
        "王斌": "139 2527 2867",
        "王少雄": "186 8387 9841",
        "苗旺旺": "138 1871 5251",
        "赵岩松": "186 1161 8385",
        "Bruce Roderick, WAINES": "186 6532 9796",
        "Oliver Viktor, RACZ": "186 1197 3165",
        "Yiftah RAUCH": "186 1045 0563",
        "尤欣": "139 1608 5072",
        "李亚民": "133 6632 0878",
        "赵镭": "138 0883 9660",
        "彭罡": "186 1263 1888",
        "Wan Leung WU": "132 6695 8816",
        "Kwan Leung WU": "132 6695 8816",
        "吴鹏": "136 1110 5901",
        "刘汇川": "188 5827 2791",
        "于龙飞": "156 5288 0812",
        "林帅": "138 1156 6711",
        "张佳妮": "136 6169 9966",
        "张欢乐": "186 1652 1529",
        "昝昭君": "182 9570 0579",
        "孙赫": "186 0102 1216",
        "He SUN": "186 0102 1216",
        "李晓龙": "138 2378 1747",
        "Xiaolong LI": "138 2378 1747",
        "李卉妍": "138 5142 0321",
        "Huiyan LI": "138 5142 0321",
        "熊立凌": "135 3821 6276",
        "Liling Xiong": "135 3821 6276",
        "HEALY, Darran William": "852 6891 2350",
        "Darran William HEALY": "852 6891 2350",
        "ROEDER, SIMONE ELKE": "852 6263 4569",
        "SIMONE ELKE ROEDER": "852 6263 4569",
        "王凯珮": "852 6858 8410",
        "HOI PUI, WONG": "852 6858 8410",
        "马坚": "189 1770 2918",
        "Jian MA": "189 1770 2918",
        "卢江": "158 0045 6521",
        "Jiang LU": "158 0045 6521",
        "孟周聪": "135 6434 5029",
        "Zhoucong Meng": "135 6434 5029",
        "张帆": "138 0135 1294",
        "Fan ZHANG": "138 0135 1294",
        "魏思远": "133 2110 4588",
        "Siyuan WEI": "133 2110 4588",
        "王晟磊": "150 2689 7493",
        "Shenglei WANG": "150 2689 7493",
        "Keith Robert, SHERREN": "137 3540 9744",
        "Keith Robert SHERREN": "137 3540 9744",
        "Rodolfo, BONETTI": "132 6284 1083",
        "Rodolfo BONETTI": "132 6284 1083",
        "危慧": "152 1349 1328",
        "Hui WEI": "152 1349 1328",
        "李辛欣": "135 5008 8666",
        "Xinxin LI": "135 5008 8666",
        "Herve Daniel, STAMM": "183 1709 0300",
        "Herve Daniel STAMM": "183 1709 0300",
        "樊婉程": "186 2017 4817",
        "Wancheng FAN": "186 2017 4817",
        "刘爽": "138 0125 8789",
        "Shuang LIU": "138 0125 8789",
        "刘凯": "135 2157 9157",
        "Kai LIU": "135 2157 9157",
        "詹佩佩": "137 1440 5925",
        "Peipei ZHAN": "137 1440 5925",
        "花佩": "186 2631 0634",
        "Pei HUA": "186 2631 0634",
        "翁英": "130 6785 2000",
        "Ying WENG": "130 6785 2000",
        "王莹": "159 1009 9069",
        "Ying WANG": "159 1009 9069",
        "程佳俊": "134 8010 3029",
        "Jiajun CHENG": "134 8010 3029",
        "蔡雨桐": "852 6426 7445",
        "Yu Tong CHOI": "852 6426 7445",
        "俞凯": "130 0579 0326",
        "Kai YU": "130 0579 0326",
        "杨杰": "186 1694 8903",
        "Jie YANG": "186 1694 8903",
        "徐卓": "139 1028 5510",
        "Zhuo XU": "139 1028 5510",
        "丁燕栒": "135 6035 3829",
        "Yanxun DING": "135 6035 3829",
        "万虹波": "133 1297 9906",
        "Hongbo WAN": "133 1297 9906",
        "王国勤": "138 1815 8715",
        "Guoqin WANG": "138 1815 8715",
        "李潇恩": "158 0599 1600",
        "Xiaoen LI": "158 0599 1600",
        "孙辉": "139 1626 9572",
        "Hui Sun": "139 1626 9572",
        "范蕾蕾": "182 1000 6866",
        "Leilei FAN": "182 1000 6866",
        "张贺新": "136 3773 1210",
        "Hexin ZHANG": "136 3773 1210",
        "李庆宏": "135 0909 0503",
        "Qinghong LI": "135 0909 0503",
        "Eduard Pascal, Roski": "49 170 1534666",
        "Eduard Pascal Roski": "49 170 1534666",
        "Peter Robert, JACKSON": "186 8245 1935",
        "Peter Robert JACKSON": "186 8245 1935",
        "廉卓群": "133 5632 3949",
        "Zhuoqun LIAN": "133 5632 3949",
        "孙浩": "136 7012 1990",
        "Hao SUN": "136 7012 1990",
        "姚艳阁": "156 0127 9399",
        "Yange YAO": "156 0127 9399",
        "茅邂文": "152 5181 7375",
        "Xiewen MAO": "152 5181 7375",
        "宋炜": "136 3256 5565",
        "Wei SONG": "136 3256 5565",
        "BEEBE, Thaddeus John": "852 6930 1609",
        "Thaddeus John BEEBE": "852 6930 1609",
        "张哲": "139 0247 5026",
        "Zhe ZHANG": "139 0247 5026",
        "李海": "136 8131 8388",
        "Hai LI": "136 8131 8388",
        "蔡国俊": "157 1220 8304",
        "朱正宇": "189 8335 3697",
        "金尚明": "136 7113 8047",
        "Shangming JIN": "136 7113 8047",
        "赵婷婷": "138 2883 3162",
        "赖小燕": "60 1239 05520",
        "Siau Mui LAI": "60 1239 05520",
        "何静文": "852 6421 0994",
        "周丽欢": "152 5710 6140",
        "AYA, MUGURUMA": "81 8071140700",
        "梁广煜": "137 9428 7177",
        "李园": "139 1178 3914",
        "孔铮": "139 1085 3981",
        "高峰": "135 8186 9017",
        "李阳": "133 6603 6567",
        "林生": "159 2162 9406",
        "谢依椿": "159 8942 4501",
        "廖关荣": "181 0755 9103",
        "林峰": "135 2260 7955",
        "丘东": "136 0044 6505",
        "王庆辉": "134 8013 9352",
        "姜磊": "135 1006 5318",
        "黄彦杰": "132 1157 2184",
        "王珍": "198 6662 9312",
        "赵国庆": "155 8849 2975",
        "王军": "853 62666900",
        "张德桃": "130 2883 6410",
        "江焰辉": "136 3148 0927",
        "孙龙": "156 9558 0691",
        "梁平": "138 2750 6225",
        "冯仁毫": "185 2028 6463",
        "焦石军": "139 2388 3525",
        "万子辰": "177 7005 7193",
        "卓辉": "157 7070 8632",
        "苏志斌": "159 0150 7150",
        "赵康": "191 6764 6172",
        "翟征宇": "134 1448 9793",
        "陈居瑜": "158 8962 6660",
        "林毅": "136 8642 0153",
        "郭春旭": "138 0136 1720",
        "Guo Chunxu": "138 0136 1720",
    }

    # ---------- 内置执照号码映射 ----------
    BUILTIN_LICENSE_MAP = {
        "吴鹏": "130103197602102115",
        "刘汇川": "330103199003191618",
        "于龙飞": "ZN00915",
        "林帅": "110227198601130015",
        "彭罡": "3240393",
        "Kwan Leung WU": "3025478",
        "Wan Leung WU": "3025478",
        "张佳妮": "10183",
        "Jiani ZHANG": "10183",
        "尤欣": "620102197604293015",
        "Xin YOU": "620102197604293015",
        "昝昭君": "14272419950203313X",
        "Zhaojun ZAN": "14272419950203313X",
        "张欢乐": "ZN00883",
        "Huanle ZHANG": "ZN00883",
        "孙赫": "4070599",
        "He SUN": "4070599",
        "李晓龙": "3781955",
        "Xiaolong LI": "3781955",
        "李卉妍": "10299",
        "Huiyan LI": "10299",
        "熊立凌": "421003199203222631",
        "Liling Xiong": "421003199203222631",
        "HEALY, Darran William": "3141079",
        "Darran William HEALY": "3141079",
        "ROEDER, SIMONE ELKE": "4213276",
        "SIMONE ELKE ROEDER": "4213276",
        "王凯珮": "10068",
        "HOI PUI, WONG": "10068",
        "马坚": "320103196607089512",
        "Jian MA": "320103196607089512",
        "卢江": "10302",
        "Jiang LU": "10302",
        "孟周聪": "10303",
        "Zhoucong Meng": "10303",
        "张帆": "2552356",
        "Fan ZHANG": "2552356",
        "魏思远": "230102198911054315",
        "Siyuan WEI": "230102198911054315",
        "王晟磊": "310109198409264012",
        "Shenglei WANG": "310109198409264012",
        "Keith Robert, SHERREN": "12262",
        "Keith Robert SHERREN": "12262",
        "Rodolfo, BONETTI": "12660",
        "Rodolfo BONETTI": "12660",
        "危慧": "10137",
        "Hui WEI": "10137",
        "李辛欣": "4209424",
        "Xinxin LI": "4209424",
        "Herve Daniel, STAMM": "2666622",
        "Herve Daniel STAMM": "2666622",
        "樊婉程": "ZN00434",
        "Wancheng FAN": "ZN00434",
        "刘爽": "4101498",
        "Shuang LIU": "4101498",
        "刘凯": "2833670",
        "Kai LIU": "2833670",
        "詹佩佩": "10283",
        "Peipei ZHAN": "10283",
        "王斌": "3340398",
        "Bin WANG": "3340398",
        "Bruce Roderick, WAINES": "3448726",
        "Bruce Roderick WAINES": "3448726",
        "花佩": "ZN00495",
        "Pei HUA": "ZN00495",
        "李亚民": "350104197107184915",
        "Yamin Li": "350104197107184915",
        "赵镭": "440301198204157271",
        "Lei ZHAO": "440301198204157271",
        "翁英": "ZN00905",
        "Ying WENG": "ZN00905",
        "王莹": "370125199004215621",
        "Ying WANG": "370125199004215621",
        "赵岩松": "410103197004017014",
        "Yansong ZHAO": "410103197004017014",
        "金尚明": "210381197511034612",
        "Shangming JIN": "210381197511034612",
        "程佳俊": "511202198208161358",
        "Jiajun CHENG": "511202198208161358",
        "张永一": "110102196605202336",
        "Yongyi ZHANG": "110102196605202336",
        "Oliver Viktor, RACZ": "000336198206158001",
        "Oliver Viktor RACZ": "000336198206158001",
        "蔡雨桐": "10269",
        "Yu Tong CHOI": "10269",
        "俞凯": "120110196912180351",
        "Kai YU": "120110196912180351",
        "杨杰": "310104198411304413",
        "Jie YANG": "310104198411304413",
        "Yiftah, RAUCH": "000972198112152001",
        "Yiftah RAUCH": "000972198112152001",
        "苗旺旺": "410781198608019797",
        "Wangwang MIAO": "410781198608019797",
        "徐卓": "110105198906307113",
        "Zhuo XU": "110105198906307113",
        "丁燕栒": "ZN00499",
        "Yanxun DING": "ZN00499",
        "万虹波": "36050219860709003X",
        "Hongbo WAN": "36050219860709003X",
        "王国勤": "2589322",
        "Guoqin WANG": "2589322",
        "李潇恩": "ZN00468",
        "Xiaoen LI": "ZN00468",
        "孙辉": "310228197810012612",
        "Hui Sun": "310228197810012612",
        "范蕾蕾": "ZN00347",
        "Leilei FAN": "ZN00347",
        "张贺新": "420106197101020435",
        "Hexin ZHANG": "420106197101020435",
        "李庆宏": "17260/1 FCL",
        "Qinghong LI": "17260/1 FCL",
        "梅峰": "17205/1 FCL",
        "Feng MEI": "17205/1 FCL",
        "Eduard Pascal, Roski": "3863535",
        "Eduard Pascal Roski": "3863535",
        "廉卓群": "ZN00430",
        "Zhuoqun LIAN": "ZN00430",
        "孙浩": "650103199102160037",
        "Hao SUN": "650103199102160037",
        "Peter Robert, JACKSON": "000044197906253001",
        "Peter Robert JACKSON": "000044197906253001",
        "王少雄": "510105198609042555",
        "Shaoxiong WANG": "510105198609042555",
        "BEEBE, Thaddeus John": "2743899",
        "Thaddeus John BEEBE": "2743899",
        "姚艳阁": "10204",
        "Yange YAO": "10204",
        "茅邂文": "ZN00903",
        "Xiewen MAO": "ZN00903",
        "宋炜": "37060219820621211X",
        "Wei SONG": "37060219820621211X",
        "张哲": "650104196604163310",
        "Zhe ZHANG": "650104196604163310",
        "李海": "110105197201106130",
        "赵婷婷": "372524198212240023",
        "李园": "110105198309149639",
        "孔铮": "11010419801116125X",
        "高峰": "130826198001175332",
        "陈居瑜": "460004197905030814",
        "林生": "350627198512262530",
        "谢依椿": "441427198601221716",
        "廖关荣": "360732199009095838",
        "王珍": "622627199605283017",
        "赵国庆": "370403200003133433",
        "王军": "1458107(8)",
        "张德桃": "360311199603192012",
        "江焰辉": "440182199307270615",
        "李阳": "11010319850705091X",
        "丘东": "11010119650406453x",
        "王庆辉": "350627198212052013",
        "姜磊": "360502198312071333",
        "黄彦杰": "450881199810196233",
        "林毅": "350102197903213219",
        "庚凡": "430104197901184015",
        "何静文": "Z630284(0)",
        "Ching Man, HO": "Z630284(0)",
        "朱正宇": "350111197207152412",
        "赖小燕": "A71366020",
        "Siau Mui LAI": "A71366020",
        "周丽欢": "330411199101195440",
        "AYA, MUGURUMA": "TT5868294",
        "梁广煜": "EK9629672",
        "林峰": "150402198501122713",
        "孙龙": "341623200010015613",
        "梁平": "441223198510246217",
        "冯仁毫": "440582199101153650",
        "焦石军": "421224198310151013",
        "卓辉": "36073219981213009X",
        "苏志斌": "350782198308221539",
        "赵康": "430321200303160170",
        "翟征宇": "140211198612050031",
        "郭春旭": "110107197305150016",
        "Guo Chunxu": "110107197305150016",
    }

    # ---------- 内置证件号码映射 ----------
    BUILTIN_ID_MAP = {
        "庚凡": "430104197901184015",
        "张永一": "110102196605202336",
        "梅峰": "510107197911242636",
        "王斌": "610104197911058331",
        "王少雄": "510105198609042555",
        "苗旺旺": "410781198608019797",
        "赵岩松": "410103197004017014",
        "Bruce Roderick, WAINES": "000336198206158001",
        "Oliver Viktor, RACZ": "000336198206158001",
        "Yiftah RAUCH": "000972198112152001",
        "尤欣": "620102197604293015",
        "李亚民": "350104197107184915",
        "赵镭": "440301198204157271",
        "彭罡": "440111198403244812",
        "Kwan Leung WU": "Z394213(A)",
        "吴鹏": "130103197602102115",
        "刘汇川": "330103199003191618",
        "于龙飞": "210103198808123928",
        "林帅": "110227198601130015",
        "张佳妮": "31010619840115002X",
        "张欢乐": "330381198705292523",
        "昝昭君": "14272419950203313X",
        "孙赫": "410102197702243012",
        "李晓龙": "420104197906150015",
        "李卉妍": "",
        "熊立凌": "421003199203222631",
        "HEALY, Darran William": "",
        "ROEDER, SIMONE ELKE": "",
        "王凯珮": "Z411582(2)",
        "马坚": "320103196607089512",
        "卢江": "420521198809280021",
        "孟周聪": "310113198307243215",
        "张帆": "211002197306050057",
        "魏思远": "230102198911054315",
        "王晟磊": "310109198409264012",
        "Keith Robert, SHERREN": "",
        "Rodolfo, BONETTI": "",
        "危慧": "43072119941115468X",
        "李辛欣": "510105198605023015",
        "Herve Daniel, STAMM": "",
        "樊婉程": "440106199608180326",
        "刘爽": "110108196308296450",
        "刘凯": "230103198103275511",
        "詹佩佩": "440301198809275123",
        "花佩": "320281198911117761",
        "翁英": "330106198801182024",
        "王莹": "370125199004215621",
        "程佳俊": "511202198208161358",
        "蔡雨桐": "V146532(5)",
        "俞凯": "120110196912180351",
        "杨杰": "310104198411304413",
        "徐卓": "110105198906307113",
        "丁燕栒": "440510199107260826",
        "万虹波": "36050219860709003X",
        "王国勤": "340822197610240215",
        "李潇恩": "510802199512100046",
        "孙辉": "310228197810012612",
        "Hui Sun": "310228197810012612",
        "范蕾蕾": "37010319861027002X",
        "张贺新": "420106197101020435",
        "李庆宏": "441402199107140256",
        "Eduard Pascal Roski": "C9TR22VZM",
        "Peter Robert JACKSON": "000044197906253001",
        "廉卓群": "370402199702018029",
        "孙浩": "650103199102160037",
        "姚艳阁": "130922199601151224",
        "茅邂文": "320683199911098627",
        "宋炜": "37060219820621211X",
        "BEEBE, Thaddeus John": "R909452(A)",
        "张哲": "650104196604163310",
        "李海": "110105197201106130",
        "赵婷婷": "372524198212240023",
        "李园": "110105198309149639",
        "孔铮": "11010419801116125X",
        "高峰": "130826198001175332",
        "陈居瑜": "460004197905030814",
        "林生": "350627198512262530",
        "谢依椿": "441427198601221716",
        "廖关荣": "360732199009095838",
        "王珍": "622627199605283017",
        "赵国庆": "370403200003133433",
        "王军": "1458107(8)",
        "张德桃": "360311199603192012",
        "江焰辉": "440182199307270615",
        "李阳": "11010319850705091X",
        "丘东": "11010119650406453x",
        "王庆辉": "350627198212052013",
        "姜磊": "360502198312071333",
        "黄彦杰": "450881199810196233",
        "林毅": "350102197903213219",
        "何静文": "Z630284(0)",
        "Ching Man, HO": "Z630284(0)",
        "朱正宇": "350111197207152412",
        "赖小燕": "A71366020",
        "Siau Mui LAI": "A71366020",
        "周丽欢": "330411199101195440",
        "AYA, MUGURUMA": "",
        "梁广煜": "",
        "林峰": "150402198501122713",
        "孙龙": "341623200010015613",
        "梁平": "441223198510246217",
        "冯仁毫": "440582199101153650",
        "焦石军": "421224198310151013",
        "卓辉": "36073219981213009X",
        "苏志斌": "350782198308221539",
        "赵康": "430321200303160170",
        "翟征宇": "140211198612050031",
        "郭春旭": "110107197305150016",
        "Guo Chunxu": "110107197305150016",
    }

    # ---------- 国籍映射 ----------
    NATION_MAP = {
        "CHN": "中国", "HKG": "香港", "DEU": "德国", "USA": "美国", "GBR": "英国",
        "FRA": "法国", "RUS": "俄罗斯", "JPN": "日本", "KOR": "韩国", "SGP": "新加坡",
        "MYS": "马来西亚", "THA": "泰国", "VNM": "越南", "PHL": "菲律宾", "IDN": "印度尼西亚",
        "IND": "印度", "AUS": "澳大利亚", "CAN": "加拿大", "BRA": "巴西", "MEX": "墨西哥",
        "ZAF": "南非", "EGY": "埃及", "NGA": "尼日利亚", "KEN": "肯尼亚", "TZA": "坦桑尼亚",
        "ZWE": "津巴布韦", "NLD": "荷兰", "ITA": "意大利", "ESP": "西班牙", "PRT": "葡萄牙",
        "GRC": "希腊", "TUR": "土耳其", "SAU": "沙特阿拉伯", "ARE": "阿联酋", "ISR": "以色列",
        "IRN": "伊朗", "PAK": "巴基斯坦", "BGD": "孟加拉", "NPL": "尼泊尔", "LKA": "斯里兰卡",
        "MMR": "缅甸", "KHM": "柬埔寨", "LAO": "老挝", "MNG": "蒙古", "PRK": "朝鲜",
        "TWN": "台湾地区", "MAC": "澳门"
    }

    AIRCRAFT_TYPE_CORRECTION = {"B3926": "LJ60"}

    def correct_aircraft_type(reg, ac_type):
        if reg in AIRCRAFT_TYPE_CORRECTION:
            corrected = AIRCRAFT_TYPE_CORRECTION[reg]
            if ac_type != corrected:
                st.info(f"✈️ 机型修正：{ac_type} → {corrected}（注册号 {reg}）")
            return corrected
        return ac_type

    def get_nation_name(code):
        code = code.strip().upper()
        return NATION_MAP.get(code, code)

    def extract_chinese_name(full_name):
        if not full_name:
            return ""
        parts = full_name.split()
        chinese_parts = [p for p in parts if re.search(r'[\u4e00-\u9fff]', p)]
        if chinese_parts:
            return " ".join(chinese_parts)
        else:
            return full_name

    def normalize_name(name):
        if not name:
            return ""
        name = re.sub(r'[\u4e00-\u9fff]+', '', name)
        name = re.sub(r'[,\s]+', ' ', name).strip()
        return name.lower()

    def find_contact(crew_name):
        if not crew_name or not BUILTIN_CONTACT_MAP:
            return ""
        chinese = extract_chinese_name(crew_name)
        if chinese and chinese in BUILTIN_CONTACT_MAP:
            return BUILTIN_CONTACT_MAP[chinese]
        norm = normalize_name(crew_name)
        if not norm:
            return ""
        for key, val in BUILTIN_CONTACT_MAP.items():
            if normalize_name(key) == norm:
                return val
        return ""

    def find_license(crew_name):
        if not crew_name or not BUILTIN_LICENSE_MAP:
            return ""
        chinese = extract_chinese_name(crew_name)
        if chinese and chinese in BUILTIN_LICENSE_MAP:
            return BUILTIN_LICENSE_MAP[chinese]
        norm = normalize_name(crew_name)
        if not norm:
            return ""
        for key, val in BUILTIN_LICENSE_MAP.items():
            if normalize_name(key) == norm:
                return val
        return ""

    def find_id(crew_name):
        if not crew_name or not BUILTIN_ID_MAP:
            return ""
        chinese = extract_chinese_name(crew_name)
        if chinese and chinese in BUILTIN_ID_MAP:
            return BUILTIN_ID_MAP[chinese]
        norm = normalize_name(crew_name)
        if norm:
            for key, val in BUILTIN_ID_MAP.items():
                if normalize_name(key) == norm:
                    return val
        return ""

    def parse_document_type(passport_no, doc_type):
        doc_type_str = str(doc_type).strip() if pd.notna(doc_type) else ""
        if doc_type_str:
            if "中华人民共和国居民身份证" in doc_type_str:
                return "身份证"
            if "港澳居民来往内地通行证" in doc_type_str:
                return "回乡证"
            return doc_type_str
        pn = str(passport_no).strip() if pd.notna(passport_no) else ""
        pn = re.sub(r'\s+', '', pn)
        if re.match(r'^[0-9]{15}$', pn) or re.match(r'^[0-9]{17}[0-9Xx]$', pn):
            return "身份证"
        else:
            return "护照"

    def safe_set_cell_value(ws, row, col, value):
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        ws.cell(row=row, column=col).value = value

    def get_value_right(ws, row, start_col):
        for col in range(start_col, start_col + 10):
            cell = ws.cell(row=row, column=col)
            if cell.value and str(cell.value).strip():
                return str(cell.value).strip()
        return ""

    def parse_utc_to_beijing(utc_str, date_str):
        try:
            time_part = utc_str.replace('Z', '').strip()
            if len(time_part) == 4:
                hour = int(time_part[:2])
                minute = int(time_part[2:])
            elif len(time_part) == 3:
                hour = int(time_part[:1])
                minute = int(time_part[1:])
            else:
                return "0000"
            day = int(re.search(r'\d+', date_str).group()) if re.search(r'\d+', date_str) else 1
            month_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                         "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
            month_str = re.search(r'[A-Za-z]{3}', date_str).group() if re.search(r'[A-Za-z]{3}', date_str) else "Jan"
            month = month_map.get(month_str[:3], 1)
            year = 2026
            dt = datetime(year, month, day, hour, minute)
            dt_beijing = dt + timedelta(hours=8)
            return dt_beijing.strftime("%H%M")
        except:
            return "0000"

    def parse_date_display(date_str):
        try:
            day = re.search(r'\d+', date_str).group()
            month_str = re.search(r'[A-Za-z]{3}', date_str).group()
            month_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                         "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
            month = month_map.get(month_str[:3], 1)
            return f"{month}月{int(day)}日"
        except:
            return date_str

    def get_beijing_date_display(utc_time_str, date_str):
        if not utc_time_str or not date_str:
            return parse_date_display(date_str)
        try:
            time_part = utc_time_str.replace('Z', '').strip()
            if len(time_part) == 4:
                hour = int(time_part[:2])
                minute = int(time_part[2:])
            elif len(time_part) == 3:
                hour = int(time_part[:1])
                minute = int(time_part[1:])
            else:
                return parse_date_display(date_str)
            day = int(re.search(r'\d+', date_str).group())
            month_str = re.search(r'[A-Za-z]{3}', date_str).group()
            month_map = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                         "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
            month = month_map.get(month_str[:3], 1)
            year = 2026
            dt_utc = datetime(year, month, day, hour, minute)
            dt_beijing = dt_utc + timedelta(hours=8)
            return f"{dt_beijing.month}月{dt_beijing.day}日"
        except:
            return parse_date_display(date_str)

    def strip_single_letter_prefix(text):
        if text and re.match(r'^[A-Za-z]\s+', text):
            return re.sub(r'^[A-Za-z]\s+', '', text)
        return text

    def parse_no_time_route(input_text, date_display):
        input_text = strip_single_letter_prefix(input_text)
        if not input_text or not input_text.strip():
            return None
        text = input_text.strip()
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        if len(lines) >= 2:
            first_line = lines[0]
            second_line = lines[1]
            flight_number = first_line.split()[0] if first_line.split() else None
            if not flight_number:
                return None
            parts = re.split(r'\s*[-—–]\s*', second_line)
            if len(parts) >= 2:
                dep_airport = parts[0].strip()
                arr_airport = parts[1].strip()
                if dep_airport and arr_airport:
                    return f"{date_display} {flight_number} {dep_airport}-{arr_airport}"
        else:
            flight_number = text.split()[0] if text.split() else None
            if not flight_number:
                return None
            time_pattern = r'(\d{1,2}:\d{2})\s*[-—–]\s*(\d{1,2}:\d{2})'
            remaining = re.sub(time_pattern, '', text).strip()
            remaining = re.sub(r'\s*\+\s*\d+\s*', '', remaining).strip()
            airport_parts = re.split(r'\s*[-—–]\s*', remaining)
            if len(airport_parts) >= 2:
                dep_airport = airport_parts[-2].strip()
                arr_airport = airport_parts[-1].strip()
                ch_dep = re.findall(r'[\u4e00-\u9fff]+', dep_airport)
                ch_arr = re.findall(r'[\u4e00-\u9fff]+', arr_airport)
                if ch_dep and ch_arr:
                    dep_airport = ''.join(ch_dep)
                    arr_airport = ''.join(ch_arr)
                elif ch_dep and not ch_arr:
                    pass
                if dep_airport and arr_airport:
                    return f"{date_display} {flight_number} {dep_airport}-{arr_airport}"
        return None

    def parse_with_time_route(input_text, date_display):
        input_text = strip_single_letter_prefix(input_text)
        if not input_text or not input_text.strip():
            return input_text
        text = input_text.strip()
        time_pattern = r'(\d{1,2}:\d{2})\s*[-—–]\s*(\d{1,2}:\d{2})'
        time_match = re.search(time_pattern, text)
        if time_match:
            dep_time = time_match.group(1).replace(':', '')
            arr_time = time_match.group(2).replace(':', '')
            remaining = re.sub(time_pattern, '', text).strip()
        else:
            time_pattern2 = r'(\d{1,2}:\d{2})\s+(\d{1,2}:\d{2})'
            time_match2 = re.search(time_pattern2, text)
            if time_match2:
                dep_time = time_match2.group(1).replace(':', '')
                arr_time = time_match2.group(2).replace(':', '')
                remaining = re.sub(time_pattern2, '', text).strip()
            else:
                return input_text
        remaining = re.sub(r'\s*\+\s*\d+\s*', '', remaining).strip()
        airport_pattern = r'(.+?)\s*[-—–]\s*(.+)'
        airport_match = re.search(airport_pattern, remaining)
        if airport_match:
            dep_airport = airport_match.group(1).strip()
            arr_airport = airport_match.group(2).strip()
            def extract_chinese(text):
                chinese = re.findall(r'[\u4e00-\u9fff]+', text)
                return ''.join(chinese) if chinese else text
            dep_airport = extract_chinese(dep_airport)
            arr_airport = extract_chinese(arr_airport)
            if dep_airport and arr_airport:
                return f"{date_display} {dep_airport}{dep_time}-{arr_time}{arr_airport}"
        return input_text

    def parse_general_declaration(file_bytes):
        wb = load_workbook(file_bytes)
        ws = wb.active
        data = {}
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if "OPERATOR:" in val:
                        data["operator"] = get_value_right(ws, cell.row, cell.column+1)
                    elif "REG NO./FLT NO.:" in val:
                        reg_val = get_value_right(ws, cell.row, cell.column+1)
                        parts = reg_val.split()
                        data["reg"] = parts[0] if parts else reg_val
                        data["flt"] = parts[0] if parts else reg_val
                        if len(parts) > 1:
                            data["flt"] = parts[1]
                    elif "AC TYPE:" in val:
                        data["ac_type_raw"] = get_value_right(ws, cell.row, cell.column+1)
                        reg = data.get("reg", "")
                        data["ac_type"] = correct_aircraft_type(reg, data["ac_type_raw"])
                    elif "FROM:" in val:
                        data["from"] = get_value_right(ws, cell.row, cell.column+1)
                    elif "TO:" in val:
                        data["to"] = get_value_right(ws, cell.row, cell.column+1)
                    elif "DATE/TIME:" in val:
                        date_time = get_value_right(ws, cell.row, cell.column+1)
                        data["date_time"] = date_time
                        if date_time:
                            parts = date_time.split()
                            data["utc_time"] = parts[0] if len(parts) > 0 else ""
                            data["date_str"] = parts[1] if len(parts) > 1 else ""
        crew_data = []
        passenger_data = []
        section = None
        for row in ws.iter_rows(min_row=1):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if "CREW MANIFEST" in val:
                        section = 'crew'
                        break
                    elif "PASSENGER MANIFEST" in val:
                        section = 'passenger'
                        break
                    elif "CARGO MANIFEST" in val or "DECLARATION OF HEALTH" in val:
                        section = None
                        break
            if section == 'crew':
                first_cell = row[0]
                if first_cell.value and isinstance(first_cell.value, (int, float)):
                    if len(row) >= 7:
                        name_cell = row[1]
                        if name_cell.value and isinstance(name_cell.value, str):
                            crew_data.append({
                                "name": name_cell.value.strip(),
                                "dob": row[2].value if row[2].value else "",
                                "gender": row[3].value if row[3].value else "",
                                "nationality": row[4].value if row[4].value else "",
                                "doc_type": row[5].value if row[5].value else "",
                                "passport_no": row[6].value if row[6].value else "",
                            })
            elif section == 'passenger':
                first_cell = row[0]
                if first_cell.value and isinstance(first_cell.value, (int, float)):
                    if len(row) >= 7:
                        name_cell = row[1]
                        if name_cell.value and isinstance(name_cell.value, str):
                            passenger_data.append({
                                "name": name_cell.value.strip(),
                                "dob": row[2].value if row[2].value else "",
                                "gender": row[3].value if row[3].value else "",
                                "nationality": row[4].value if row[4].value else "",
                                "doc_type": row[5].value if row[5].value else "",
                                "passport_no": row[6].value if row[6].value else "",
                            })
        return data, crew_data, passenger_data

    def fill_template(template_bytes, data, crew_list, passenger_list, route_display):
        try:
            wb = load_workbook(template_bytes)
        except Exception as e:
            if "Bad magic number" in str(e) or "BadZipFile" in str(e):
                st.error("❌ 模板文件格式不正确。请确保模板为 **.xlsx** 格式（非 .xls）。")
                st.info("💡 解决方法：用 Excel 打开该模板，选择“另存为”，将文件类型选为 **Excel工作簿（.xlsx）**，然后重新上传。")
            raise e

        ws = wb.active

        if not passenger_list:
            for row in ws.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "飞行目的" in cell.value:
                        target_row = cell.row + 1
                        safe_set_cell_value(ws, target_row, 2, "调机")
                        break
                else:
                    continue
                break

        info_row = None
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if val in ["机型", "注册号", "航班号", "航班行程"]:
                        info_row = cell.row
                        break
            if info_row:
                break

        if info_row:
            data_row = info_row + 1
            safe_set_cell_value(ws, data_row, 2, data.get("ac_type", ""))
            safe_set_cell_value(ws, data_row, 3, data.get("reg", ""))
            safe_set_cell_value(ws, data_row, 4, data.get("flt", ""))
            safe_set_cell_value(ws, data_row, 5, route_display if route_display else "")

        # 机长
        if len(crew_list) >= 1:
            crew = crew_list[0]
            for row in ws.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "机长" in cell.value:
                        row_num = cell.row
                        safe_set_cell_value(ws, row_num, 2, extract_chinese_name(crew["name"]))
                        safe_set_cell_value(ws, row_num, 3, crew.get("gender", ""))
                        safe_set_cell_value(ws, row_num, 4, crew.get("dob", ""))
                        id_num = find_id(crew["name"])
                        if id_num:
                            safe_set_cell_value(ws, row_num, 5, id_num)
                        else:
                            safe_set_cell_value(ws, row_num, 5, crew.get("passport_no", ""))
                        license_num = find_license(crew["name"])
                        safe_set_cell_value(ws, row_num, 6, license_num)
                        contact = find_contact(crew["name"])
                        safe_set_cell_value(ws, row_num, 7, contact)
                        break
                else:
                    continue
                break

        # 副驾驶
        if len(crew_list) >= 2:
            crew = crew_list[1]
            for row in ws.iter_rows(min_row=1, max_row=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "副驾驶" in cell.value:
                        row_num = cell.row
                        safe_set_cell_value(ws, row_num, 2, extract_chinese_name(crew["name"]))
                        safe_set_cell_value(ws, row_num, 3, crew.get("gender", ""))
                        safe_set_cell_value(ws, row_num, 4, crew.get("dob", ""))
                        id_num = find_id(crew["name"])
                        if id_num:
                            safe_set_cell_value(ws, row_num, 5, id_num)
                        else:
                            safe_set_cell_value(ws, row_num, 5, crew.get("passport_no", ""))
                        license_num = find_license(crew["name"])
                        safe_set_cell_value(ws, row_num, 6, license_num)
                        contact = find_contact(crew["name"])
                        safe_set_cell_value(ws, row_num, 7, contact)
                        break
                else:
                    continue
                break

        cabin_crew = None
        mechanic = None
        for i in range(2, len(crew_list)):
            crew = crew_list[i]
            gender = str(crew.get("gender", "")).strip()
            if gender in ["女", "Female", "F"] and cabin_crew is None:
                cabin_crew = crew
            elif gender in ["男", "Male", "M"] and mechanic is None:
                mechanic = crew
            if cabin_crew and mechanic:
                break

        # 乘务行
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "乘务" in cell.value:
                    row_num = cell.row
                    if cabin_crew:
                        safe_set_cell_value(ws, row_num, 2, extract_chinese_name(cabin_crew["name"]))
                        safe_set_cell_value(ws, row_num, 3, cabin_crew.get("gender", ""))
                        safe_set_cell_value(ws, row_num, 4, cabin_crew.get("dob", ""))
                        id_num = find_id(cabin_crew["name"])
                        if id_num:
                            safe_set_cell_value(ws, row_num, 5, id_num)
                        else:
                            safe_set_cell_value(ws, row_num, 5, cabin_crew.get("passport_no", ""))
                        license_num = find_id(cabin_crew["name"]) or find_license(cabin_crew["name"])
                        safe_set_cell_value(ws, row_num, 6, license_num)
                        contact = find_contact(cabin_crew["name"])
                        safe_set_cell_value(ws, row_num, 7, contact)
                    else:
                        safe_set_cell_value(ws, row_num, 2, "无")
                        safe_set_cell_value(ws, row_num, 3, "")
                        safe_set_cell_value(ws, row_num, 4, "")
                        safe_set_cell_value(ws, row_num, 5, "")
                        safe_set_cell_value(ws, row_num, 6, "")
                        safe_set_cell_value(ws, row_num, 7, "")
                    break
            else:
                continue
            break

        # 机务行
        for row in ws.iter_rows(min_row=1, max_row=50):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "机务" in cell.value:
                    row_num = cell.row
                    if mechanic:
                        safe_set_cell_value(ws, row_num, 2, extract_chinese_name(mechanic["name"]))
                        safe_set_cell_value(ws, row_num, 3, mechanic.get("gender", ""))
                        safe_set_cell_value(ws, row_num, 4, mechanic.get("dob", ""))
                        id_num = find_id(mechanic["name"])
                        if id_num:
                            safe_set_cell_value(ws, row_num, 5, id_num)
                        else:
                            safe_set_cell_value(ws, row_num, 5, mechanic.get("passport_no", ""))
                        license_num = find_license(mechanic["name"])
                        safe_set_cell_value(ws, row_num, 6, license_num)
                        contact = find_contact(mechanic["name"])
                        safe_set_cell_value(ws, row_num, 7, contact)
                    else:
                        safe_set_cell_value(ws, row_num, 2, "无")
                        safe_set_cell_value(ws, row_num, 3, "")
                        safe_set_cell_value(ws, row_num, 4, "")
                        safe_set_cell_value(ws, row_num, 5, "")
                        safe_set_cell_value(ws, row_num, 6, "")
                        safe_set_cell_value(ws, row_num, 7, "")
                    break
            else:
                continue
            break

        passenger_start_row = None
        for row in ws.iter_rows(min_row=1, max_row=100):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip()
                    if "姓名" in val and "性别" in val and "出生日期" in val:
                        passenger_start_row = cell.row + 1
                        break
            if passenger_start_row:
                break

        if passenger_start_row is None:
            for row in ws.iter_rows(min_row=1, max_row=100):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "乘客信息" in cell.value:
                        passenger_start_row = cell.row + 2
                        break
                if passenger_start_row:
                    break

        if passenger_start_row:
            for i, pax in enumerate(passenger_list):
                row_num = passenger_start_row + i
                for col in range(1, 7):
                    safe_set_cell_value(ws, row_num, col, None)
                safe_set_cell_value(ws, row_num, 1, extract_chinese_name(pax["name"]))
                safe_set_cell_value(ws, row_num, 2, pax.get("gender", ""))
                safe_set_cell_value(ws, row_num, 3, pax.get("dob", ""))
                safe_set_cell_value(ws, row_num, 4, get_nation_name(pax.get("nationality", "")))
                doc_type = pax.get("doc_type", "")
                if pd.notna(doc_type) and str(doc_type).strip():
                    doc_type_clean = parse_document_type("", doc_type)
                else:
                    doc_type_clean = parse_document_type(pax.get("passport_no", ""), "")
                safe_set_cell_value(ws, row_num, 5, doc_type_clean)
                safe_set_cell_value(ws, row_num, 6, pax.get("passport_no", ""))

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ---------- 功能1 UI ----------
    st.subheader("📂 上传文件")
    st.info("⚠️ 注意：模板文件必须是 **.xlsx** 格式（非 .xls）。联系方式、执照号码及证件号码已内置，无需额外上传。")

    data_file = st.file_uploader("上传 GD单（General Declaration）Excel（.xlsx）", type=["xlsx"], key="data")
    template_file = st.file_uploader("上传总调模板：Jetops申请一览-", type=["xlsx"], key="template")

    if data_file and template_file:
        try:
            data, crew_list, passenger_list = parse_general_declaration(data_file)
            st.success(f"✅ 解析成功：机组 {len(crew_list)} 人，乘客 {len(passenger_list)} 人")

            if crew_list:
                st.write("提取的机组信息：", pd.DataFrame(crew_list))
            if passenger_list:
                st.write("提取的乘客信息（前5行）：", pd.DataFrame(passenger_list).head(5))

            from_code = data.get("from", "")
            to_code = data.get("to", "")
            date_str = data.get("date_str", "")
            utc_time = data.get("utc_time", "")
            default_route = ""
            date_display = get_beijing_date_display(utc_time, date_str) if date_str else ""
            if date_str and from_code and to_code:
                if utc_time:
                    bj_time = parse_utc_to_beijing(utc_time, date_str)
                else:
                    bj_time = "0000"
                default_route = f"{date_display} {from_code} {bj_time} XXXX {to_code}"
            else:
                default_route = f"{from_code}-{to_code}" if from_code and to_code else ""

            st.subheader("📋 提取的航班信息")
            st.info(f"✈️ 默认航班行程：{default_route}")

            raw_route = st.text_input("📝 自定义航班行程 / 下载文件名", value=default_route).strip()

            with_time = parse_with_time_route(raw_route, date_display)
            if with_time != raw_route and with_time is not None:
                route_display = with_time
            else:
                route_display = raw_route if raw_route else default_route

            no_time = parse_no_time_route(raw_route, date_display)
            if no_time is not None:
                file_name_base = no_time
            else:
                reg = data.get("reg", "")
                if reg and from_code and to_code:
                    file_name_base = f"{date_display} {reg} {from_code}-{to_code}"
                else:
                    file_name_base = route_display

            safe_file_name = re.sub(r'[\\/*?:"<>|]', "_", file_name_base).strip()
            if not safe_file_name:
                safe_file_name = "备案表"
            download_file_name = f"{safe_file_name}.xlsx"

            result_bytes = fill_template(template_file, data, crew_list, passenger_list, route_display)

            st.download_button(
                label="⬇️ 下载填充后的备案表",
                data=result_bytes,
                file_name=download_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ 处理失败：{e}")
            st.exception(e)
    else:
        st.info("👆 请同时上传 GD单 和 模板文件。")


# ================================================================
# 功能2：世界时行程（带记忆对比功能）
# ================================================================
with tab2:
    st.markdown("从Jetops系统导出的北京时间的行程 Excel 文件转换为世界时的行程，便于复制粘贴。")
    st.info("💡 每次上传将自动与上一次记录对比，新增或变更的航段会在下方红色高亮显示。")

    # ---------- 历史数据管理 ----------
    HISTORY_FILE = "flight_history.json"

    def load_history():
        if os.path.exists(HISTORY_FILE):
            try:
                with open(HISTORY_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                return {"records": []}
        else:
            return {"records": []}

    def save_history(history):
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)

    def parse_time_column(val):
        if pd.isna(val):
            return None
        if isinstance(val, (pd.Timestamp, datetime)):
            return val.strftime('%H:%M')
        if hasattr(val, 'strftime'):
            return val.strftime('%H:%M')
        s = str(val).strip()
        if ':' in s:
            return s[:5]
        return s

    def convert_to_utc(date_val, time_str):
        if pd.isna(date_val) or time_str is None:
            return None
        if isinstance(date_val, (pd.Timestamp, datetime)):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            date_str = str(date_val).split()[0]
        dt_str = f"{date_str} {time_str}"
        try:
            dt_local = pd.to_datetime(dt_str)
            dt_utc = dt_local - timedelta(hours=8)
            return dt_utc
        except:
            return None

    def format_utc(dt):
        if dt is None:
            return ""
        months = ['JAN','FEB','MAR','APR','MAY','JUN',
                  'JUL','AUG','SEP','OCT','NOV','DEC']
        day = dt.day
        month = months[dt.month - 1]
        hour = dt.hour
        minute = dt.minute
        return f"{day:02d}{month} {hour:02d}{minute:02d}Z"

    def generate_plans(df):
        required = ['飞机注册号', '出发地', '到达地', '出发日期', '计划出发', '到达日期', '预计到达', '用途']
        for col in required:
            if col not in df.columns:
                st.error(f"❌ 缺少列：{col}")
                return None

        df = df.dropna(subset=['出发地', '到达地', '出发日期', '计划出发'])
        if df.empty:
            st.warning("没有有效的航段数据")
            return None

        plans = {}
        for idx, row in df.iterrows():
            reg = row['飞机注册号']
            if pd.isna(reg) or str(reg).strip() == '':
                reg = "N/A"
            else:
                reg = str(reg).strip()

            dep_time = parse_time_column(row['计划出发'])
            arr_time = parse_time_column(row['预计到达'])
            if dep_time is None or arr_time is None:
                continue

            dep_utc = convert_to_utc(row['出发日期'], dep_time)
            arr_utc = convert_to_utc(row['到达日期'], arr_time)
            if dep_utc is None or arr_utc is None:
                continue

            use = str(row['用途']) if not pd.isna(row['用途']) else ''
            flight_type = 'FERRY' if '调机' in use else 'PAX'

            line = (f"ETD {row['出发地']} {format_utc(dep_utc)} // "
                    f"ETA {row['到达地']} {format_utc(arr_utc)}  {flight_type}")

            if reg not in plans:
                plans[reg] = []
            plans[reg].append((dep_utc, line))

        result = {}
        for reg, items in plans.items():
            items.sort(key=lambda x: x[0])
            lines = [reg]
            lines.extend([item[1] for item in items])
            result[reg] = "\n".join(lines)

        return result

    def sort_plans(plans_dict):
        priority_order = ['B652Q', 'B65AP', 'B652S', 'MLLIN', 'N88AY', 'B652R']
        all_keys = list(plans_dict.keys())
        priority_keys = [k for k in priority_order if k in all_keys]
        remaining_keys = [k for k in all_keys if k not in priority_order and k != "N/A"]
        remaining_keys.sort()
        na_keys = [k for k in all_keys if k == "N/A"]
        sorted_keys = priority_keys + remaining_keys + na_keys
        return {k: plans_dict[k] for k in sorted_keys}

    def diff_plans(old_plans, new_plans):
        changes = {}
        all_regs = set(old_plans.keys()) | set(new_plans.keys())
        for reg in all_regs:
            old_lines = set(old_plans.get(reg, "").split('\n')) if old_plans.get(reg) else set()
            new_lines = set(new_plans.get(reg, "").split('\n')) if new_plans.get(reg) else set()
            old_lines.discard(reg)
            new_lines.discard(reg)
            added = new_lines - old_lines
            for line in added:
                changes[(reg, line)] = 'added'
        return changes

    # ---------- UI ----------
    uploaded_file_2 = st.file_uploader("📤 上传航段数据导出（北京时间）", type=["xlsx"], key="worldtime")

    if uploaded_file_2 is not None:
        try:
            df = pd.read_excel(uploaded_file_2, skiprows=1)
            st.success("✅ 文件读取成功")

            new_plans = generate_plans(df)
            if new_plans is None:
                st.stop()

            sorted_new_plans = sort_plans(new_plans)

            history = load_history()
            old_plans = {}
            if history["records"]:
                last_record = history["records"][-1]
                old_plans = last_record.get("data", {})

            changes = diff_plans(old_plans, new_plans)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            filename = uploaded_file_2.name
            new_record = {
                "timestamp": timestamp,
                "filename": filename,
                "data": new_plans
            }
            history["records"].append(new_record)
            if len(history["records"]) > 20:
                history["records"] = history["records"][-20:]
            save_history(history)

            st.subheader("📋 生成的飞行计划（红色为新增/变更）")

            # 每个注册号独立显示
            for reg, text in sorted_new_plans.items():
                lines = text.split('\n')
                has_changes = any((reg, line) in changes for line in lines if line != reg)

                if has_changes:
                    st.markdown(f"**✈️ {reg}** 🔴 <span style='color:red;font-size:0.9rem;'>（有新增或变更）</span>", unsafe_allow_html=True)
                else:
                    st.markdown(f"**✈️ {reg}**")

                plain_lines = []
                for line in lines:
                    if line == reg:
                        continue
                    plain_lines.append(line)

                # 显示纯文本代码框（自带复制按钮）
                plain_text = "\n".join(plain_lines)
                st.code(plain_text, language="text")

            # 全部合并的纯文本（备用）
            full_text = ""
            for reg, text in sorted_new_plans.items():
                full_text += f"{text}\n\n"
            with st.expander("📦 全部计划合并（点击展开）"):
                st.code(full_text, language="text")

            # ---------- 历史记录 ----------
            with st.expander("📜 查看历史上传记录"):
                if history["records"]:
                    for i, rec in enumerate(history["records"]):
                        st.write(f"{i+1}. {rec['timestamp']} - {rec['filename']}")
                else:
                    st.write("暂无历史记录")

            if st.button("🗑️ 清除所有历史记录", key="clear_history"):
                save_history({"records": []})
                st.success("历史已清除，请刷新页面")
                st.rerun()

        except Exception as e:
            st.error(f"❌ 处理出错：{e}")
            st.stop()
    else:
        st.info("请上传一个符合格式的 Excel 文件。")

    st.markdown("---")
    st.caption("🛠️ 工具说明：日期/时间按北京时间（UTC+8）自动转换为世界时（Z）。对比功能基于上一次上传的记录。")


# ================================================================
# 功能3：航路处理工具
# ================================================================
with tab3:
    st.markdown("支持表格格式（带N/E坐标）/中文描述格式，自动精简航路+添加#前缀，兼容不规整数据")

    if "last_processed_input_route" not in st.session_state:
        st.session_state.last_processed_input_route = ""
    if "result_text_route" not in st.session_state:
        st.session_state.result_text_route = ""

    def parse_coord(coord_str):
        letter = coord_str[0]
        num_part = coord_str[1:]
        if letter == 'N':
            deg = int(num_part[0:2])
            minute = int(num_part[2:4])
            sec_part = num_part[4:]
            if '.' in sec_part:
                sec_float = float(sec_part)
                sec_int = int(round(sec_float))
            else:
                sec_int = int(sec_part)
            if sec_int >= 60:
                sec_int -= 60
                minute += 1
                if minute >= 60:
                    minute -= 60
                    deg += 1
            return f"{deg:02d}{minute:02d}{sec_int:02d}"
        elif letter == 'E':
            deg = int(num_part[0:3])
            minute = int(num_part[3:5])
            sec_part = num_part[5:]
            if '.' in sec_part:
                sec_float = float(sec_part)
                sec_int = int(round(sec_float))
            else:
                sec_int = int(sec_part)
            if sec_int >= 60:
                sec_int -= 60
                minute += 1
                if minute >= 60:
                    minute -= 60
                    deg += 1
            return f"{deg:03d}{minute:02d}{sec_int:02d}"
        else:
            raise ValueError(f"未知的坐标前缀: {letter}")

    def base_name(s):
        return s.split('@')[0]

    def is_open_point(s):
        base = base_name(s)
        if re.match(r'^[A-Z]{2,5}$', base):
            return True
        if re.match(r'^P[A-Z]+$', base):
            return True
        return False

    def is_p_point(s):
        base = base_name(s)
        return re.match(r'^P\d+$', base) is not None

    def clean_route(r):
        if r.startswith('#'):
            return r[1:]
        return r

    def is_open_route(rt):
        return rt and rt[0] not in ('H', 'J', 'V')

    def extract_table(text):
        tokens = text.strip().split()
        start_idx = 0
        for i, tok in enumerate(tokens):
            if tok.isdigit() and 1 <= int(tok) <= 40:
                start_idx = i
                break
        tokens = tokens[start_idx:]

        lines = []
        i = 0
        while i < len(tokens):
            if tokens[i].isdigit():
                line = [tokens[i]]
                i += 1
                while i < len(tokens) and not tokens[i].isdigit():
                    line.append(tokens[i])
                    i += 1
                lines.append(line)

        points = []
        routes = []
        for line in lines:
            lat_idx = None
            for idx, tok in enumerate(line):
                if tok.startswith('N') and tok[1:].replace('.', '', 1).isdigit():
                    lat_idx = idx
                    break
            if lat_idx is None:
                continue
            lon_idx = lat_idx + 1
            if lon_idx >= len(line) or not line[lon_idx].startswith('E'):
                continue
            lat_str = line[lat_idx]
            lon_str = line[lon_idx]

            route = None
            if lon_idx + 1 < len(line):
                next_tok = line[lon_idx + 1]
                if re.match(r'^[A-Z][A-Z0-9]*$', next_tok) and not next_tok[0].isdigit():
                    route = next_tok

            point_name = None
            for j in range(lat_idx - 1, 0, -1):
                tok = line[j]
                if is_open_point(tok) or is_p_point(tok):
                    point_name = tok
                    break
            if point_name is None:
                continue

            if is_p_point(point_name):
                lat_int = parse_coord(lat_str)
                lon_int = parse_coord(lon_str)
                point_display = f"{point_name}@{lat_int}N{lon_int}E"
            else:
                point_display = point_name

            points.append(point_display)
            if route is not None:
                routes.append(route)

        seq = []
        for i in range(len(points)):
            seq.append(points[i])
            if i < len(routes):
                seq.append(routes[i])
        return seq

    def extract_chinese(text):
        text = re.sub(r'[\u4e00-\u9fa5，、。；：""''（）【】]', ' ', text)
        words = text.split()
        seq = []
        for w in words:
            if '(' in w and ')' in w:
                m = re.search(r'\(([A-Z]+)\)', w)
                if m:
                    point = m.group(1)
                    prefix = w[:w.find('(')]
                    m_route = re.search(r'([A-Z]\d+)$', prefix)
                    if m_route:
                        seq.append(m_route.group(1))
                    seq.append(point)
            elif re.match(r'^[A-Z]\d+[A-Z]{2,5}$', w) or re.match(r'^[A-Z]\d+P\d+$', w):
                m = re.match(r'^([A-Z]\d+)([A-Z]{2,5}|P\d+)$', w)
                if m:
                    seq.append(m.group(1))
                    seq.append(m.group(2))
            elif re.match(r'^[A-Z]\d+$', w):
                seq.append(w)
            elif is_open_point(w) or is_p_point(w):
                seq.append(w)
        return seq

    def step1_extract(text):
        if re.search(r'N\d{5,6}(?:\.\d+)?\s+E\d{6,7}(?:\.\d+)?', text):
            return extract_table(text), 'table'
        else:
            return extract_chinese(text), 'chinese'

    def step2_reduce(seq):
        L = seq[:]
        changed = True
        while changed:
            changed = False
            n = len(L)
            candidates = []
            for i in range(0, n, 2):
                if not is_open_point(L[i]):
                    continue
                if i + 1 >= n:
                    continue
                first_route = clean_route(L[i+1])
                if not is_open_route(first_route):
                    continue
                for j in range(i+2, n, 2):
                    all_same = True
                    for k in range(i+1, j, 2):
                        rt = clean_route(L[k])
                        if rt != first_route or not is_open_route(rt):
                            all_same = False
                            break
                    if not all_same:
                        break
                    if is_open_point(L[j]):
                        length = (j - i) // 2
                        if length >= 2:
                            candidates.append((i, j, length))
            if not candidates:
                break
            candidates.sort(key=lambda x: -x[2])
            best_i, best_j, _ = candidates[0]
            new_segment = [L[best_i], L[best_i+1], L[best_j]]
            L = L[:best_i] + new_segment + L[best_j+1:]
            changed = True
        return L

    def step3_add_hash(seq):
        pts = seq[0::2]
        rts = seq[1::2]
        m = len(rts)

        def is_closed_route(rt):
            return rt.startswith(('H', 'J', 'V'))

        def is_p(pt):
            base = base_name(pt)
            return re.match(r'^P\d+$', base) is not None

        res = [pts[0]]
        for i, rt in enumerate(rts):
            left = pts[i]
            right = pts[i+1]
            need_hash = False
            if is_closed_route(rt):
                need_hash = True
            elif is_p(left) or is_p(right):
                need_hash = True
            res.append('#' + rt if need_hash else rt)
            res.append(right)
        return res

    # ---------- 功能3 UI ----------
    st.markdown("""
        <style>
        .stButton>button {border-radius: 8px; height: 2.5rem; font-size: 1rem;}
        .stProgress>div>div {background-color: #1890ff;}
        .stCaption {color: #666666; font-size: 0.9rem;}
        </style>
    """, unsafe_allow_html=True)

    input_text_route = st.text_area(
        "📋 请输入待处理的航路文本",
        key="input_text_route",
        height=300,
        placeholder="粘贴民航航线数据，支持多行表格格式/纯中文描述格式..."
    )

    btn_col1, btn_col2, btn_col3 = st.columns([2, 2, 8])
    with btn_col1:
        process_btn = st.button("⚙️ 处理", type="primary", use_container_width=True, key="process_route")
    with btn_col2:
        clear_btn = st.button("🗑️ 清空", use_container_width=True, key="clear_route")

    if clear_btn:
        st.session_state.input_text_route = ""
        st.session_state.last_processed_input_route = ""
        st.session_state.result_text_route = ""
        st.rerun()

    if process_btn and st.session_state.get("input_text_route", "").strip():
        progress_bar = st.progress(0)
        status_text = st.empty()
        total_steps = 4
        current_step = 0

        try:
            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（识别输入类型）")
            seq, fmt = step1_extract(st.session_state.input_text_route)

            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（精简相同开放航路）")
            if fmt == 'table':
                seq = step2_reduce(seq)

            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（添加航路#前缀）")
            if fmt == 'table':
                seq = step3_add_hash(seq)

            current_step += 1
            progress_bar.progress(current_step / total_steps)
            status_text.text(f"处理中：第{current_step}步/共{total_steps}步（生成最终结果）")
            result = ' '.join(seq) if seq else "⚠️ 未提取到有效航路数据"

            st.session_state.result_text_route = result
            st.session_state.last_processed_input_route = st.session_state.input_text_route

            progress_bar.empty()
            status_text.empty()
            st.success("✅ 处理完成！结果如下：")

        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error(f"❌ 处理失败：{str(e)}")
            with st.expander("🔍 查看详细错误信息", expanded=False):
                st.code(traceback.format_exc(), language="text")

    if st.session_state.get("result_text_route", ""):
        current_input = st.session_state.get("input_text_route", "")
        last_input = st.session_state.last_processed_input_route

        st.subheader("📊 处理结果", divider="blue")

        if current_input != last_input:
            st.warning("⚠️ 输入已更改，当前显示的是上一次处理的结果，如需更新请点击「处理」按钮。")

        st.code(st.session_state.result_text_route, language="text")

    if not st.session_state.get("result_text_route", "") and not st.session_state.get("input_text_route", "").strip():
        st.info("💡 提示：粘贴航路数据后，点击「处理」即可，支持30+行不规整表格数据")

    st.markdown("---")
    st.caption("✈️ 支持表格格式（带N/E坐标）/中文描述格式，自动精简航路+添加#前缀")
